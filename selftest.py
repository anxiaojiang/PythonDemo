import os
import subprocess
import win32api
import win32con
from version import Version
import shutil
from pywinauto.application import Application
from pywinauto import Desktop
from pywinauto import keyboard
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import shutil
import xlwt
import xlrd
from xlutils.copy import copy
import openpyxl
from openpyxl.styles import  PatternFill
from openpyxl.styles import  Border,Side

def copyTestData():
    process = subprocess.Popen(r'\\mlangfs1\public\v-xingsh\F12dll\copyF12TestAppx.cmd', 
        stdout=subprocess.PIPE, stdin=subprocess.PIPE)
    for line in iter(process.stdout.readline, ''):
        print(line.rstrip())
        if(line == b'run finish\r\n'):
            break
    process.kill()

def getVersion():
    key = win32api.RegOpenKey(win32con.HKEY_LOCAL_MACHINE, 
        r'SOFTWARE\Microsoft\Windows NT\CurrentVersion', 0, win32con.KEY_READ)
    builds = win32api.RegQueryValueEx(key, 'BuildLabEx')[0]
    builds = builds.split('.')
    version = Version(builds[3], '{0}.{1}.{2}'.format(builds[0], builds[1], builds[4]), builds[2])
    return version
    
def copyMstrcaFile():
    version = getVersion()
    filename = r'\\winbuilds\release\{0}\{1}\{2}\bin\AppxTools\Certificates\TestRoot\MSTRCA2010.cer'.format(
            version.branch, version.build, version.cpu)
    shutil.copyfile(filename, r'C:\\MSTRCA2010.cer')

def registerCertFile():
    process = subprocess.Popen(r'certutil.exe -addstore root C:\\MSTRCA2010.cer', 
        stdout=subprocess.PIPE, stdin=subprocess.PIPE)
    for line in iter(process.stdout.readline, ''):
        print(line.rstrip())
        if(line == b'CertUtil: -addstore command completed successfully.\r\n'):
            break
    process.kill()

def openDevelopmentMode():
    runProgram('ms-settings:developers')
    settings = Desktop(backend="uia")["Settings"]
    btn = settings['For developersGroupBox'].RadioButton3
    btn.click()
    settings['Use developer featuresDialog'].Yes.click()
    settings.close()
    
def runEdgeDevTools():
    os.system(r'C:\Windows\SystemApps\Microsoft.MicrosoftEdge_8wekyb3d8bbwe\MicrosoftEdgeDevToolsClientBeta.appx')
    install = Desktop(backend="uia")["App Installer"]
    install['Install Microsoft Edge DevTools Client? Install'].click()
    time.sleep(60)
    install.CloseButton.click()

def runProgram(program):
    keyboard.send_keys('{VK_LWIN down}''{r down}''{VK_LWIN up}')
    run = Desktop(backend="win32")["Run"]
    run.Edit.type_keys(program)
    run.OK.click()

def runEdge():
    driver = webdriver.Edge()
    driver.get(r'https://ie-snap/scratchtests/devtools/SelfTest')
    keyboard.send_keys('{F12}')
    frame = driver.find_element_by_id('selftest-runner')
    driver.switch_to.frame(frame)
    time.sleep(20)
    btn = driver.find_element_by_id('startRun')
    btn.click()
    while True:
        last = driver.find_element_by_xpath("//div[@class='testRow'][last()]/span[contains(@class, 'status')]")
        clazz = last.get_attribute("class")
        if 'status-failed' in clazz or 'status-passed' in clazz or 'status-timedout' in clazz:
            break
        else:
            time.sleep(10)
    testrows = driver.find_elements_by_xpath("//div[@class='testRow']")
    version = getVersion()
    excelfile = r"\\mlangfs1\public\F12Team\SelfTestForBranch\{0}\{0}.xlsx".format(version.branch)
    if not checkFileExists(excelfile):
        createResultFile(excelfile, testrows)
    index = getExcelIndex(excelfile)
    excel = openpyxl.load_workbook(excelfile)
    sheet = excel.get_sheet_by_name(u'All')
    i = 2;
    passfill = PatternFill("solid", fgColor="FFFFFF")
    failfill = PatternFill("solid", fgColor="FF0000")
    border = Border(left=Side(style='thin',color='000000'),
        right=Side(style='thin',color='000000'),
        top=Side(style='thin',color='000000'),
        bottom=Side(style='thin',color='000000'))
    strtime = time.strftime("%m%d", time.localtime())
    if strtime != sheet.cell(1, 6).value:
        sheet.insert_cols(6)
        sheet.cell(1, 6).value = strtime
        sheet.cell(1, 6).border = border
    i = 2;
    for testrow in testrows:
        elem = testrow.find_element_by_class_name('priority')
        priority = elem.get_attribute("innerHTML")[1:]
        testname = testrow.find_element_by_class_name('name').get_attribute("innerHTML")
        pos = 0
        try:
            pos = index.index(testname)
        except ValueError as e:
            pos = i
            index.insert(i, testname)
            sheet.insert_rows(i)
            names = testname.split('.')
            feature = names[0]
            area = names[1]
            name = names[2]
            sheet.cell(i, 1).value = priority
            sheet.cell(i, 1).border = border
            sheet.cell(i, 2).value = feature
            sheet.cell(i, 2).border = border
            sheet.cell(i, 3).value = area
            sheet.cell(i, 3).border = border
            sheet.cell(i, 4).value = name
            sheet.cell(i, 4).border = border
        clazz = testrow.find_element_by_class_name('status').get_attribute("class")
        sheet.cell(pos, 6).border = border
        if 'status-passed' in clazz:
            sheet.cell(pos, 6).value = "Pass"
            sheet.cell(pos, 6).fill = passfill
        elif 'status-failed' in clazz or 'status-timedout' in clazz:
            sheet.cell(pos, 6).value = "Failed"
            sheet.cell(pos, 6).fill = failfill
        else:
            sheet.cell(pos, 6).value = ""
            sheet.cell(pos, 6).fill = passfill
        i+=1
    excel.save(excelfile)
        
def getExcelIndex(file):
    excel = openpyxl.load_workbook(file)
    sheet = excel.get_sheet_by_name(u'All')
    index = ['', '']
    i = 2;
    while True:
        if sheet.cell(i, 4).value != None:
            index.append('{0}.{1}.{2}'.format(
                sheet.cell(i, 2).value, 
                sheet.cell(i, 3).value, 
                sheet.cell(i, 4).value))
            i+=1
        else:
            break
    return index

def checkFileExists(file):
    dir = os.path.dirname(file)
    if not os.path.exists(dir):
        os.makedirs(dir)
    if not os.path.exists(file):
        shutil.copy(os.path.dirname(dir) + r'\template.xlsx', file)
        return False
    return True

def createResultFile(file, testrows):
    excel = openpyxl.load_workbook(file)
    sheet = excel.get_sheet_by_name(u'All')
    i = 2
    border = Border(left=Side(style='thin',color='000000'),
        right=Side(style='thin',color='000000'),
        top=Side(style='thin',color='000000'),
        bottom=Side(style='thin',color='000000'))
    for testrow in testrows:
        elem = testrow.find_element_by_class_name('priority')
        priority = elem.get_attribute("innerHTML")[1:]
        testname = testrow.find_element_by_class_name('name').get_attribute("innerHTML")
        names = testname.split('.')
        feature = names[0]
        area = names[1]
        name = names[2]
        sheet.cell(i, 1).value = priority
        sheet.cell(i, 1).border = border
        sheet.cell(i, 2).value = feature
        sheet.cell(i, 2).border = border
        sheet.cell(i, 3).value = area
        sheet.cell(i, 3).border = border
        sheet.cell(i, 4).value = name
        sheet.cell(i, 4).border = border
        sheet.cell(i, 5).border = border
        i+=1
    excel.save(file);

def installEdgeDriver():
    process = subprocess.Popen(r'DISM.exe /Online /Add-Capability /CapabilityName:Microsoft.WebDriver~~~~0.0.1.0', 
        stdout=subprocess.PIPE, stdin=subprocess.PIPE)
    for line in iter(process.stdout.readline, ''):
        print(line.rstrip())
        if(line == b'The operation completed successfully.\r\n'):
            break
    process.kill()


copyTestData()
copyMstrcaFile() 
registerCertFile()
openDevelopmentMode()
runEdgeDevTools()
runEdge()
print('success')



    
